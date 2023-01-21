import zipfile
from datetime import timedelta
import pandas as pd
from openpyxl import load_workbook
import warnings
from settings_RSS import Settings
import data_manipulation_module as dmm
import xlsxwriter

settings = Settings()
warnings.filterwarnings('ignore', '.*data validation.*', )


# todo create a class for a client directory

def check_excel_for_duplicates(filename, sheetname, column):
    """
    This function aims to read an excel file and alerts the user if there are duplicate clients in the dataframe.
    :param filename: file name or path
    :return:
    """
    print(f"Checking {filename} for duplicates!")
    df = pd.read_excel(filename,sheet_name=sheetname)
    while True:
        try:
            boolean = not df[column].is_unique
            break
        except KeyError:
            try:
                column = "Name"
            except:
                column = input(f"There is no column named '{column}' in this file:\n {filename}, please enter a different column name: ")

    if boolean:
        print(f"\nThere are duplicate accounts in\nfile:{filename}\nsheet: {sheetname}\nPlease investigate.")

def write_excel(filename, sheetname, dataframe):
    """
    Write a dataframe to a specific excel sheet.
    :param filename: filename/path of excel file.
    :param sheetname: name of the excel sheet you want to write to.
    :param dataframe: the dataframe that will be written to the excel sheet
    :return:
    """
    while True:
        try:
            book = load_workbook(filename)
            writer = pd.ExcelWriter(filename, engine='openpyxl')
            writer.book = book

            writer.sheets = {ws.title: ws for ws in book.worksheets}

            dataframe.to_excel(writer, sheetname, index=False)

            writer.save()

            break
        except zipfile.BadZipfile:
            run = input(f"There is a problem with the file: {filename}. Please replace this file and enter 'run': ")
            if run == 'run':
                continue
        except PermissionError:
            run = input("Looks like the file you are trying to access is already open. Please close the file and enter"
                        " 'run' to continue: ")
            if run == 'run':
                continue
        except FileNotFoundError:
            print("File Doesn't Exist Yet, Let Me Make It!")
            workbook = xlsxwriter.Workbook(filename)
            workbook.add_worksheet('Data')
            workbook.close()
            print(f"{filename} created!")


# noinspection PyTypeChecker
def update_coach_directories(dataframe):
    """
    Updates the client directory file for each set of coach initials in settings
    :param dataframe: reporting_df will be used to update client directories
    :return:
    """

    for i in range(len(settings.coach_initials)):
        write_excel(
            settings.coach_directories[i], 'Data',
            dataframe[
                (dataframe['Case Manager'] == f"{settings.coach_initials[i]}") &
                (pd.to_datetime(dataframe['Original Intake Date']) >= (pd.Timestamp('today') - timedelta(days=730)))])

    if settings.username == "Fred":
        for i in range(len(settings.coach_initials)):
            write_excel(
                settings.od_coach_directories[i], 'Data',
                dataframe[
                    (dataframe['Case Manager'] == f"{settings.coach_initials[i]}") &
                    (pd.to_datetime(dataframe['Original Intake Date']) >= (
                                pd.Timestamp('today') - timedelta(days=730)))])

def update_complete_client_directory(df):
    """
    Update the complete_client_directory_xlsx.
    :param df:
    :return:
    """
    directory_df = df[pd.to_datetime(df['Original Intake Date'], errors='coerce') >= pd.to_datetime('2019-07-01')]
    write_excel(settings.complete_client_directory, 'Data', directory_df)

    if settings.username == "Fred":
        write_excel(settings.od_complete_client_directory, 'Data', directory_df)



def update_client_directories():
    """
    Updates all client directories.
    :return:
    """
    df = pd.read_csv(settings.reporting_df_csv)
    df = dmm.coerce_to_datetime(df,'Original Intake Date','Cohort Date')
    update_complete_client_directory(df)
    update_coach_directories(df)
