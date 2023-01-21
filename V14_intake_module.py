import json
import os
from pathlib import Path
import PyPDF2
import pandas as pd
import requests
from openpyxl import load_workbook
from settings_RSS import Settings

settings = Settings()


def extract_pdf_data(pdf):
    """
    Pull data from the custom fields of a fillable pdf and organize it in a dictionary.
    :param pdf: PDF file with fillable text fields.
    :return: A messy dictionary of all data contained in PDF.
    """
    while True:
        try:
            with open(pdf, 'rb') as new_intake:
                read_pdf = PyPDF2.PdfFileReader(new_intake)
                messy_dict = read_pdf.getFields()
                print(f"Extracting data from {pdf}!")
            break
        except FileNotFoundError:
            print(f"{pdf} is not in this directory!")
            pdf = input('Enter the filepath of a pdf that is in this directory: ')
    return messy_dict


def clean_pdf_data(messy_dict):
    """
    Clean and format dictionary built from pdf custom fields.
    :param messy_dict: Messy dictionary taken from PDF file.
    :return: A cleanly formatted dictionary containing  data from the PDF fields.
    """
    clean_dict = {}
    for field_name, value in messy_dict.items():
        field_value = value.get('/V', None)
        if field_name not in clean_dict:
            clean_dict[field_name] = []
        clean_dict[field_name].append(field_value)
    return {k: str(v[0]) for k, v in dict.items(clean_dict)}


def remove_punctuation(dictionary, *punctuation):
    """
    Remove unnecessary punctuation from values.
    :param dictionary: Organized PDF data.
    :param punctuation: The punctuation(s) to be removed from fields.
    :return: Modified dictionary without specified punctuation.
    """
    for p in punctuation:
        for key, value in dictionary.items():
            dictionary[key] = value.replace(f"{p}", "")
    return dictionary


def format_name_fields(new_dict):
    """
    Combine and format all name fields to return one full properly formatted name.
    :param new_dict: Dictionary of PDF fields
    :return: Modified dictionary featuring a properly formatted 'Name' key.
    """
    if (new_dict['Suffix'] == 'None') & (new_dict['MI'] != 'None'):
        new_dict['name'] = new_dict['Last Name'].rstrip() + ", " + new_dict['First Name'].rstrip() + " " + \
                           new_dict['MI'].rstrip() + "."

    elif (new_dict['Suffix'] != 'None') & (new_dict['MI'] == 'None'):
        new_dict['name'] = new_dict['Last Name'].rstrip() + " " + new_dict['Suffix'].rstrip() + ", " + \
                           new_dict['First Name'].rstrip()

    elif (new_dict['Suffix'] == 'None') & (new_dict['MI'] == 'None'):
        new_dict['name'] = new_dict['Last Name'].rstrip() + ", " + new_dict['First Name'].rstrip()

    else:
        new_dict['name'] = new_dict['Last Name'].rstrip() + " " + new_dict['Suffix'].rstrip() + ", " + \
                           new_dict['First Name'].rstrip() + " " + new_dict['MI'].rstrip() + "."

    new_dict['name'] = new_dict['name'].title()
    return new_dict


def format_multiple_choice(new_dict, choices, new_field_name):
    """
    Check the values for all individual income sources and format a single income_sources field.
    :param new_dict: Dictionary made from PDF fields.
    :param choices: List of potential choices from checkbox fields.
    :param new_field_name: Name of new dictionary key displaying all selected multiple choice items.
    :return: Modified dictionary featuring new key for the multiple choice category.
    """
    new_dict[f"{new_field_name}"] = []

    for i in choices:
        if new_dict[f"{i}"] == '/Yes':
            new_dict[f"{new_field_name}"].append(f"{i}")
        del new_dict[f'{i}']
    new_dict[f"{new_field_name}"] = ', '.join(new_dict[f"{new_field_name}"])

    return new_dict


def format_true_false_checkboxes(new_dict, field_name, value='Yes', else_value='No'):
    """
    Format checkbox fields to return Y/N values. For exceptions assign unique value= or else_value= parameters.
    :param new_dict: Dictionary of PDF fields.
    :param field_name: Name of dictionary key to apply function.
    :param value: Yes or No. Default='Yes'
    :param else_value: Yes or No. Default='No'.
    :return: Modified dictionary with processed Y/N key.
    """
    new_dict[field_name] = value if new_dict[field_name] == '/Yes' else else_value
    return new_dict


def format_income_category(new_dict, prefix, income_category):
    """
    Read income category checkboxes to determine income category.
    :param new_dict: Dictionary of PDF fields.
    :param prefix: Specified prefix to add to checkboxes.
    :param income_category: The income category specified by series of checkboxes.
    :return: Modified dictionary featuring readable income category keys.
    """
    checkbox_list = [f'{prefix}1', f'{prefix}2', f'{prefix}3', f'{prefix}4', f'{prefix}5', f'{prefix}6', f'{prefix}7',
                     f'{prefix}8']

    for i in checkbox_list:
        if new_dict[f'{i}'] == '/Yes':
            new_dict['incomecategory'] = income_category
        del new_dict[f"{i}"]

    return new_dict


def check_state_field(new_dict):
    """
    Ensure that the state field has been filled out properly.
    :param new_dict: Dictionary of PDF fields
    :return: Modified dictionary that ensures state field isn't empty.
    """
    if (
            (new_dict['addresscounty/state'] == "") |
            (new_dict['addresscounty/state'] == "None") |
            (new_dict['addresscounty/state'] == " ")
    ):
        new_dict['addresscounty/state'] = "MD"

    return new_dict


def check_source_field(new_dict):
    """
    Ensure that the source field has been filled out properly.
    :param new_dict: Dictionary of PDF fields.
    :return: Modified dictionary with properly formatted 'source' key.
    """
    if (
            (new_dict['source'] == "") |
            (new_dict['source'] == "None") |
            (new_dict['source'] == " ")
    ):
        new_dict['source'] = "None Entered"

    return new_dict


def job_coach_formatter(new_dict):
    """
    Match job coach initials with with id number.
    :param new_dict: Dictionary of PDF fields.
    :return: Modified dictionary with an accurate ownerid key.
    """
    owners = ['CH', 'ML', 'RB', 'RV', 'SH', 'MM', 'MS','VF']
    owner_ids = ['36971', '36972', '36974', '31833', '20431', '36975', '36976', '36978']

    for i in range(len(owners)):
        if new_dict['owner'] == f'{owners[i]}':
            new_dict['ownerid'] = f'{owner_ids[i]}'
            break
        else:
            new_dict['ownerid'] = '36974'
    return new_dict


def remove_null_blank_values(new_dict):
    """
    Remove null or blank values from the dictionary.
    :param new_dict: Dictionary of PDF fields
    :return: Modified dictionary that doesn't contain null or 'empty' values.
    """
    null_values = ['None', 'none', '', ' ']

    for i in null_values:
        for key, value in dict(new_dict).items():
            if value == f"{i}":
                del new_dict[key]
    return new_dict


def check_returning_client(new_dict):
    """
    Checks whether the intake is for a returning client.
    :param new_dict: Dictionary of PDF fields.
    :return: True if returning_client field indicated positive.
    """
    if new_dict['returning_client'] == '/Yes':
        return True


def assign_program_of_interest(new_dict):
    """
    Assigns a value to the program of interest category.
    :param new_dict: Dictionary of PDF fields
    :return: Modified dictionary that properly formats training program field for posting.
    """
    new_dict['programofinterest'] = new_dict['originaltrainingprogram']
    return new_dict


def create_account(new_dict):
    """
    Create account by posting dictionary to RSS.
    :param new_dict: Dictionary of PDF fields
    :return: Response from POST request.
    """
    print(len(new_dict))
    payload = json.dumps(new_dict)
    print(payload)
    headers = settings.posting_headers
    response = requests.post(settings.accounts_url, headers=headers, data=payload)
    print(response)
    print(response.status_code)
    print(response.reason)
    if response.status_code != 201:
        check = input(f"\nIntake for {new_dict['name']} could not be processed!\nPlease investigate the status code and "
                      f"intake packet to identify problem. \nEnter: 'run' when ready to continue: ")

    return response


def patch_account(new_dict):
    """
    Patch intake document for returning client to existing account.
    :param new_dict: Dictionary of PDF fields.
    :return:
    """
    # Check for existing account.
    parameters = {'name': new_dict['name']}
    parameters = json.dumps(parameters)
    url = f"{settings.accounts_url}?q={parameters}"
    headers = settings.posting_headers
    r = requests.get(url, headers=headers)
    # If account exists find accountid and create new account
    if r.json()['metadata']['total_count'] != 0:
        account_id = r.json()['list'][0]['record']['id']
        url = f"{settings.documents_url}/{account_id}"
        payload = json.dumps(new_dict)
        response = requests.patch(url, headers=headers, data=payload)
        if response.status_code == 201:
            print(f"{response} - {new_dict['name']} Account Patched!")
        else:
            print(f"{new_dict['name']} Account could not be patched. We'll try creating a new account.")
            create_account(new_dict)
            print(f"New account created for {new_dict['name']}, check "
                  f"RSS for redundancy.")
    else:
        # Create New account
        create_account(new_dict)
        print(f"New account created for {new_dict['name']}, check "
              f"RSS for redundancy.")


def dictionary_to_dataframe(dictionary):
    """
    Convert a dictionary to a dataframe.
    :param dictionary: Dictionary of PDF fields
    :return: Dataframe made from dictionary.
    """
    return (
        pd.DataFrame(list(dictionary.items()), columns=['keys', 'values'])
        .set_index('keys')
        .T
    )


def add_df_to_next_xlsx_row(file_path, df):    # todo create excel file for intake data
    """
    Write dataframe to an excel sheet in the next available row.
    :param file_path: Filepath to intake data xlsx file.
    :param df: Dataframe made from PDF dictionary.
    :return: Dataframe made from PDF dictionary.
    """
    while True:
        try:
            book = load_workbook(file_path)
            writer = pd.ExcelWriter(file_path, engine='openpyxl')
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            df.to_excel(
                writer, startrow=writer.sheets['Sheet1'].max_row,
                index=False, header=False)
            writer.save()
            writer.close()
            print('Dataframe successfully written to the excel file!')
            break
        except PermissionError:
            print("Excel workbook must be closed before you can "
                  "write data to it.")
            input("Close workbook and enter 'r' to try again: ")
        except FileNotFoundError:
            print("File cannot be found!")
            file_path = input("Please enter the correct filepath: ")
    return df

def format_affected_by_covid(new_dict):
    if new_dict['affectedbycovid'] == 'Yes':
        format_multiple_choice(new_dict, settings.covid_effects, 'affectedbycovid')
    return new_dict

def format_incumbent_details(new_dict):
    if new_dict['howlongjobless'] == 'Still Working':
        new_dict['incumbentdetails'] = ''.join(
            f"{new_dict[f'{i}']}, " for i in settings.incumbent_details)
        new_dict['formeremploymentdetails'] = "Incumbent Worker"
    else:
        new_dict['incumbentdetails'] = "Not Incumbent"
        new_dict['formeremploymentdetails'] = ''.join(
            f"{new_dict[f'{i}']}, " for i in settings.incumbent_details)
        new_dict['formeremploymentdetails'] += f", {new_dict['end_date']}"
    for i in settings.incumbent_details:
        del new_dict[f'{i}']
    del new_dict['end_date']
    return new_dict




def process_digital_intake(pdf):
    """
    Combine the previously defined functions to process a digital intake by storing data in an excel file and posting
    the data to a new account on RSS.
    :param pdf: Digital intake document.
    :return:
    """
    messy_dict = extract_pdf_data(pdf)
    new_dict = clean_pdf_data(messy_dict)
    new_dict = remove_punctuation(new_dict, ',')
    new_dict = format_name_fields(new_dict)
    new_dict = format_multiple_choice(new_dict, settings.income_sources, 'incomesources')
    new_dict = format_multiple_choice(new_dict, settings.races, 'race')
    new_dict = format_multiple_choice(new_dict, settings.barriers, 'barriers')
    new_dict = format_true_false_checkboxes(new_dict, 'latino')
    new_dict = format_true_false_checkboxes(new_dict, 'femalehh')
    new_dict['children'] = new_dict['in custody']
    new_dict = format_income_category(new_dict, 'ELI', 'Extremely Low Income')
    new_dict = format_income_category(new_dict, 'LI', 'Low Income')
    new_dict = format_income_category(new_dict, 'MI', 'Moderate Income')
    new_dict = format_income_category(new_dict, 'O', 'Over Median Income')
    new_dict = check_state_field(new_dict)
    new_dict = job_coach_formatter(new_dict)
    new_dict = assign_program_of_interest(new_dict)
    new_dict = format_incumbent_details(new_dict)
    new_dict = format_multiple_choice(new_dict, settings.incumbent_benefits,'incumbentbenefits')
    new_dict = format_true_false_checkboxes(new_dict, 'returningclient')
    new_dict['scheduledtraining'] = new_dict['scheduledtraining_af_date']
    del new_dict['scheduledtraining_af_date']

    new_dict = format_affected_by_covid(new_dict)

    len_new_dict = len(new_dict)

    for i in settings.v14_rem_list:
        del new_dict[f"{i}"]
    new_dict = remove_null_blank_values(new_dict)
    print(new_dict)
    r = create_account(new_dict)
    print(f"Account for {new_dict['name']} has been created!")


def iterate_through_intakes(folder_path, new_folder_path, extension):
    """
    Iterate through each intake in directory.
    :param folder_path: Directory where intakes are downloaded.
    :param new_folder_path: processed_document path
    :return:
    """
    for file in os.listdir(folder_path):
        filename = os.fsdecode(file)
        try:
            if filename.endswith(f"{extension}"):
                pdf = os.path.join(folder_path, filename)
                if extension == ".pdf":
                    process_digital_intake(pdf)
                new_location = Path(new_folder_path, filename)
                try:
                    os.replace(pdf, new_location)
                    print(f"{filename} has been relocated.")
                except PermissionError:
                    print("Permission Error, please relocate file "
                          "manually.")
            continue
        except AttributeError:
            print(f"{filename} is not an accepted intake document. "
                  f"An account cannot be created. \
                We will still send it to the 'processed_intakes' "
                  f"folder.")
            pdf = os.path.join(folder_path, filename)
            new_location = Path(new_folder_path, filename)
            #os.replace(pdf, new_location)
            #print(f"{filename} has been relocated.")
        except KeyError:
            print(f"{filename} resulted in a Key Error. PLEASE INVESTIGATE "
                  f"An account cannot be created. \
                We will still send it to the 'processed_intakes' "
                  f"folder.")
            pdf = os.path.join(folder_path, filename)
            new_location = Path(new_folder_path, filename)
            #os.replace(pdf, new_location)
            #print(f"{filename} has been relocated.")


def download_documents(auth, document_ids):
    """
    Downloads documents from RSS.
    :param auth: RSS authorization.
    :param document_ids: ID number of documents to be downloaded.
    :return:
    """
    for i in document_ids:
        try:
        #for i in document_ids:
            # get information about the document.
            r = requests.get(f'https://apiv4.reallysimplesystems.com/documents/{i}/content',
                             headers=auth)
            print(f"{r} - Document {i} Downloaded")
            # identify document's content-disposition.
            cd = r.headers.get('Content-Disposition')
            # if document name is contained in content-disposition file format name
            if cd.find('filename'):
                filename = cd.rsplit('=', 1)[1].replace('"', '')
           # else:
            #    filename = 'unknown'
            # create document featuring formatted file name and document content
            open(filename, 'wb').write(r.content)
        except AttributeError:
            # If there's an attribute error, print an error message.
            print("We couldn't find a filename for document id.")
            r = requests.get(f'https://apiv4.reallysimplesystems.com/documents/{i}/content',
                             headers=auth)
            open("unknown", 'wb').write(r.content)


def populate_document_ids(auth):
    """
    Identify the document ID numbers that you would like to download.
    :param auth: RSS authorization
    :return: List of document ID's for download.
    """
    # create an empty list to store ids and set starting page
    id_list = []
    page = 1
    while True:
        # request page of data from rss
        url = f"{settings.documents_url}{settings.record_limit}&page={page}"
        r = requests.get(url, headers=auth)

        for record in r.json()['list']:
            # identify record id
            record_id = record['record']['id']
            # add record id to the end of id_list
            id_list.append(record_id)
        # prepare to request next page
        page += 1
        # break loop if there is no more data
        if not r.json()['metadata']['has_more']:
            break
    return id_list


def delete_documents_from_rss(auth, document_ids):
    """
    Delete the documents that you got from the New Clients Account for RSS.
    :param auth: RSS authorization
    :param document_ids: List of document ID's to be deleted.
    :return:
    """
    for i in document_ids:
        r = requests.delete(
            f'https://apiv4.reallysimplesystems.com/documents/{i}',
            headers=auth)
        print(f"{r} - Document ID: {i} Deleted from RSS.")


def complete_intake_process(auth, file_path, new_file_path):
    """
    Prompts the entry of document ID's for download, downloads documents, uploads intake document data to RSS creating
    new accounts, and moves documents to the processed intake folder.
    :param auth: RSS authorization.
    :param file_path: Directory where files will be downloaded.
    :param new_file_path: processed_documents filepath.
    :return:
    """
    documents = populate_document_ids(settings.fred_authorization)
    download_documents(auth, documents)
    for i in settings.file_extensions:
        iterate_through_intakes(file_path, new_file_path, i)
        print(f"Sorting {i} files.")
    delete_documents_from_rss(auth, documents)
    print("Intake Process Complete!")
